from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Avg
from .models import Group, Presenter, Evaluator, Score, WeightConfig
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SESSION_KEYS = {'tutor': 'tutor_auth', 'invited': 'invited_auth', 'dashboard': 'dashboard_auth'}

CRITERIA = [
    ('content', 'Content'),
    ('presentation_skills', 'Presentation Skills'),
    ('time_management', 'Time Management'),
    ('language', 'Language'),
    ('questions_answers', 'Questions & Answers'),
]

def is_authenticated(request, role):
    return request.session.get(SESSION_KEYS.get(role)) is True

def parse_score(raw):
    """Accept 2,5 or 2.5 → float, clamp 0-4, round to nearest 0.5."""
    if raw is None or str(raw).strip() == '':
        return None
    try:
        v = float(str(raw).replace(',', '.'))
        v = max(0.0, min(4.0, v))
        return round(round(v * 2) / 2, 1)  # round to nearest 0.5
    except (ValueError, TypeError):
        return None

def build_rankings():
    groups = Group.objects.prefetch_related('presenters').all()
    ranked = []
    for g in groups:
        ws = g.weighted_final_score()
        ranked.append({
            'group': g,
            'presenters': g.get_presenter_names(),
            'weighted_score': ws,
            'student_avg': g.average_total_by_type('student'),
            'tutor_avg': g.average_total_by_type('tutor'),
            'invited_avg': g.average_total_by_type('invited'),
            'breakdown': g.score_breakdown(),
            'score_count': g.scores.count(),
        })
    ranked.sort(key=lambda x: (x['weighted_score'] or 0), reverse=True)
    for i, r in enumerate(ranked, 1):
        r['rank'] = i
    return ranked


# ── Dashboard (tutor-only) ────────────────────────────────────────────────────

def dashboard(request):
    if not is_authenticated(request, 'dashboard'):
        return redirect('dashboard_password')
    rankings = build_rankings()
    config = WeightConfig.get_active()
    all_scores = Score.objects.all()
    criteria_impact = None
    if all_scores.exists():
        agg = all_scores.aggregate(
            content=Avg('content'), presentation_skills=Avg('presentation_skills'),
            time_management=Avg('time_management'), language=Avg('language'),
            questions_answers=Avg('questions_answers'),
        )
        total_avg = sum(v for v in agg.values() if v)
        if total_avg:
            criteria_impact = {
                'Content': round((agg['content'] or 0) / total_avg * 100, 1),
                'Presentation Skills': round((agg['presentation_skills'] or 0) / total_avg * 100, 1),
                'Time Management': round((agg['time_management'] or 0) / total_avg * 100, 1),
                'Language': round((agg['language'] or 0) / total_avg * 100, 1),
                'Q&A': round((agg['questions_answers'] or 0) / total_avg * 100, 1),
            }
    evaluator_counts = {et: Evaluator.objects.filter(evaluator_type=et).count()
                        for et in ['student', 'tutor', 'invited']}
    return render(request, 'evaluations/dashboard.html', {
        'rankings': rankings, 'config': config,
        'criteria_impact': criteria_impact,
        'evaluator_counts': evaluator_counts,
        'total_scores': all_scores.count(),
    })


def dashboard_password(request):
    config = WeightConfig.get_active()
    error = None
    if request.method == 'POST':
        entered = request.POST.get('password', '').strip()
        if config and entered == config.tutor_password:
            request.session['dashboard_auth'] = True
            return redirect('dashboard')
        else:
            error = "Incorrect password."
    return render(request, 'evaluations/dashboard_password.html', {'error': error})


def group_detail(request, pk):
    if not is_authenticated(request, 'dashboard'):
        return redirect('dashboard_password')
    group = get_object_or_404(Group, pk=pk)
    breakdown = group.score_breakdown()
    scores = group.scores.select_related('evaluator').order_by('evaluator__evaluator_type')
    config = WeightConfig.get_active()
    ws = group.weighted_final_score()
    type_impact = None
    if config and ws:
        total_w = float(config.total_weight())
        type_impact = {}
        for label, avg, weight in [
            ('Students', group.average_total_by_type('student'), float(config.student_weight)),
            ('Tutors', group.average_total_by_type('tutor'), float(config.tutor_weight)),
            ('Invited', group.average_total_by_type('invited'), float(config.invited_weight)),
        ]:
            if avg is not None:
                contribution = avg * weight / total_w
                type_impact[label] = {
                    'avg': avg, 'weight': weight,
                    'contribution': round(contribution, 2),
                    'pct_of_final': round(contribution / ws * 100, 1),
                }
    return render(request, 'evaluations/group_detail.html', {
        'group': group, 'breakdown': breakdown, 'scores': scores,
        'config': config, 'type_impact': type_impact, 'weighted_score': ws,
        'criteria': CRITERIA,
    })


def rankings_view(request):
    if not is_authenticated(request, 'dashboard'):
        return redirect('dashboard_password')
    return render(request, 'evaluations/rankings.html', {
        'ranked': build_rankings(), 'config': WeightConfig.get_active(),
    })


# ── Evaluation forms ──────────────────────────────────────────────────────────

def eval_landing(request):
    return render(request, 'evaluations/eval_landing.html', {'config': WeightConfig.get_active()})


def eval_form(request, role):
    if role not in ('student', 'tutor', 'invited'):
        return redirect('eval_landing')
    config = WeightConfig.get_active()
    if role in ('tutor', 'invited') and not is_authenticated(request, role):
        return redirect('eval_password', role=role)

    # Check if this browser already submitted for this role
    session_key = f'submitted_{role}'
    if request.session.get(session_key):
        submitted_name = request.session.get(f'submitted_{role}_name', '')
        return render(request, 'evaluations/eval_form.html', {
            'role': role,
            'role_label': {'student': 'Student', 'tutor': 'Tutor', 'invited': 'Invited Evaluator'}[role],
            'already_submitted': True,
            'submitted_name': submitted_name,
        })

    groups = Group.objects.prefetch_related('presenters').all()

    if request.method == 'POST':
        evaluator_name = request.POST.get('evaluator_name', '').strip()
        if not evaluator_name:
            messages.error(request, "Please enter your name.")
            return redirect('eval_form', role=role)

        # DB-level check: if this name already submitted scores, block
        existing_evaluator = Evaluator.objects.filter(
            name__iexact=evaluator_name, evaluator_type=role
        ).first()
        if existing_evaluator and existing_evaluator.scores.exists():
            request.session[session_key] = True
            request.session[f'submitted_{role}_name'] = evaluator_name
            messages.warning(request, f"An evaluation was already submitted under the name '{evaluator_name}'. Each person can only submit once.")
            return redirect('eval_form', role=role)

        evaluator, _ = Evaluator.objects.get_or_create(name=evaluator_name, evaluator_type=role)
        saved, skipped, errors = 0, 0, []

        for group in groups:
            prefix = f"group_{group.pk}_"
            try:
                content = parse_score(request.POST.get(prefix + 'content'))
                pres    = parse_score(request.POST.get(prefix + 'presentation_skills'))
                time_m  = parse_score(request.POST.get(prefix + 'time_management'))
                lang    = parse_score(request.POST.get(prefix + 'language'))
                qa      = parse_score(request.POST.get(prefix + 'questions_answers', ''))

                # require at least the 4 main scores
                if any(v is None for v in [content, pres, time_m, lang]):
                    errors.append(f"{group.name}: missing required scores")
                    continue

                if Score.objects.filter(group=group, evaluator=evaluator).exists():
                    skipped += 1
                    continue

                Score.objects.create(
                    group=group, evaluator=evaluator,
                    content=content, presentation_skills=pres,
                    time_management=time_m, language=lang,
                    questions_answers=qa,
                )
                saved += 1
            except Exception as e:
                errors.append(f"{group.name}: {e}")

        if saved or skipped:
            request.session[session_key] = True
            request.session[f'submitted_{role}_name'] = evaluator_name

        result_messages = []
        if saved:
            result_messages.append(f"{saved} score(s) saved. Thank you, {evaluator_name}!")
        if skipped:
            result_messages.append(f"{skipped} group(s) already scored by you -- skipped.")
        if errors:
            result_messages.append("Issues: " + "; ".join(errors))

        return render(request, 'evaluations/eval_form.html', {
            'role': role,
            'role_label': {'student': 'Student', 'tutor': 'Tutor', 'invited': 'Invited Evaluator'}[role],
            'just_submitted': True,
            'result_messages': result_messages,
            'submitted_name': evaluator_name,
            'had_errors': bool(errors and not saved),
        })

    return render(request, 'evaluations/eval_form.html', {
        'groups': groups, 'role': role,
        'role_label': {'student': 'Student', 'tutor': 'Tutor', 'invited': 'Invited Evaluator'}[role],
        'config': config,
        'criteria': CRITERIA,
    })


def eval_password(request, role):
    if role not in ('tutor', 'invited'):
        return redirect('eval_landing')
    config = WeightConfig.get_active()
    error = None
    if request.method == 'POST':
        entered = request.POST.get('password', '').strip()
        expected = config.tutor_password if role == 'tutor' else config.invited_password
        if config and entered == expected:
            request.session[SESSION_KEYS[role]] = True
            return redirect('eval_form', role=role)
        else:
            error = "Incorrect password. Please try again."
    label = 'Tutor' if role == 'tutor' else 'Invited Evaluator'
    return render(request, 'evaluations/eval_password.html', {'role': role, 'label': label, 'error': error})


# ── Admin clear data ──────────────────────────────────────────────────────────

def clear_data(request):
    """Admin-only: wipe all submitted scores and evaluators."""
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('dashboard_password')
    if request.method == 'POST':
        confirm = request.POST.get('confirm', '')
        if confirm == 'DELETE':
            Score.objects.all().delete()
            Evaluator.objects.all().delete()
            messages.success(request, "All scores and evaluator records have been cleared.")
            return redirect('dashboard')
        else:
            messages.error(request, "Incorrect confirmation. Type DELETE to confirm.")
    return render(request, 'evaluations/clear_data.html')


# ── Excel export ──────────────────────────────────────────────────────────────

def export_excel(request):
    if not is_authenticated(request, 'dashboard'):
        return redirect('dashboard_password')

    wb = openpyxl.Workbook()
    config = WeightConfig.get_active()

    hf  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    bf  = Font(name='Arial', bold=True, size=10)
    nf  = Font(name='Arial', size=10)
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left', vertical='center')
    thin = Side(style='thin', color='D1D5DB')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    C_STU='3B82F6'; C_TUT='10B981'; C_INV='8B5CF6'
    C_FIN='4F46E5'; C_HDR='1F2937'; C_SUB='374151'

    def hdr(cell, text, bg):
        cell.value=text; cell.font=hf
        cell.fill=PatternFill('solid', start_color=bg)
        cell.alignment=ctr; cell.border=bdr

    # ── Sheet 1: Individual Scores ──
    ws1 = wb.active; ws1.title="Individual Scores"; ws1.freeze_panes='A3'
    ws1.merge_cells('A1:M1')
    t=ws1['A1']; t.value="Individual Evaluation Scores"
    t.font=Font(name='Arial', bold=True, size=13, color='FFFFFF')
    t.alignment=ctr; t.fill=PatternFill('solid', start_color=C_HDR)
    ws1.row_dimensions[1].height=30; ws1.row_dimensions[2].height=35

    hdrs=['Group','Presenters','Evaluator Name','Evaluator Type',
          'Content (/4)','Presentation (/4)','Time (/4)','Language (/4)',
          'Q&A (/4)','Total (/20)','% Score','Submitted At']
    widths=[16,28,22,18,12,14,12,12,10,12,10,20]
    for c,(h,w) in enumerate(zip(hdrs,widths),1):
        hdr(ws1.cell(row=2,column=c),h,C_SUB)
        ws1.column_dimensions[get_column_letter(c)].width=w

    TYPE_BG={'student':'DBEAFE','tutor':'D1FAE5','invited':'EDE9FE'}
    scores=Score.objects.select_related('group','evaluator').order_by(
        'group__name','evaluator__evaluator_type','evaluator__name')
    for row,s in enumerate(scores,3):
        bg=PatternFill('solid', start_color=TYPE_BG.get(s.evaluator.evaluator_type,'FFFFFF'))
        vals=[s.group.name, s.group.get_presenter_names(), s.evaluator.name,
              s.evaluator.get_evaluator_type_display(),
              s.content, s.presentation_skills, s.time_management, s.language,
              s.questions_answers if s.questions_answers is not None else '',
              s.total(), f"{s.percentage()}%",
              s.submitted_at.strftime('%Y-%m-%d %H:%M') if s.submitted_at else '']
        for c,v in enumerate(vals,1):
            cell=ws1.cell(row=row,column=c,value=v)
            cell.font=nf; cell.alignment=ctr if c in (1,2,3,4,10,11) else lft
            cell.border=bdr; cell.fill=bg

    # ── Sheet 2: Averages by Type ──
    ws2=wb.create_sheet("Averages by Type"); ws2.freeze_panes='A4'
    ws2.merge_cells('A1:P1')
    t2=ws2['A1']; t2.value="Score Averages by Evaluator Type"
    t2.font=Font(name='Arial',bold=True,size=13,color='FFFFFF')
    t2.alignment=ctr; t2.fill=PatternFill('solid',start_color=C_HDR)
    ws2.row_dimensions[1].height=30

    sub_hdrs=['Content','Presentation','Time','Language','Q&A','Avg Total','# Eval']
    for cell in [ws2.cell(row=2,column=1),ws2.cell(row=2,column=2)]:
        cell.font=hf; cell.fill=PatternFill('solid',start_color=C_HDR)
        cell.alignment=ctr; cell.border=bdr
    ws2.cell(row=2,column=1).value='Group'
    ws2.cell(row=2,column=2).value='Presenters'
    col=3
    for label,color in [('STUDENTS',C_STU),('TUTORS',C_TUT),('INVITED EVALUATORS',C_INV)]:
        ws2.merge_cells(start_row=2,start_column=col,end_row=2,end_column=col+6)
        mc=ws2.cell(row=2,column=col)
        mc.value=label; mc.font=hf
        mc.fill=PatternFill('solid',start_color=color)
        mc.alignment=ctr; mc.border=bdr; col+=7
    ws2.merge_cells(start_row=2,start_column=col,end_row=2,end_column=col+1)
    mc=ws2.cell(row=2,column=col); mc.value='FINAL'; mc.font=hf
    mc.fill=PatternFill('solid',start_color=C_FIN); mc.alignment=ctr; mc.border=bdr

    sub_all=['Group','Presenters']+sub_hdrs+sub_hdrs+sub_hdrs+['Weighted Final (/20)','% Score']
    sub_colors=[C_HDR,C_HDR]+[C_STU]*7+[C_TUT]*7+[C_INV]*7+[C_FIN]*2
    for c,(h,sc) in enumerate(zip(sub_all,sub_colors),1):
        hdr(ws2.cell(row=3,column=c),h,sc)
        ws2.column_dimensions[get_column_letter(c)].width=13
    ws2.column_dimensions['A'].width=16; ws2.column_dimensions['B'].width=28

    groups=Group.objects.prefetch_related('presenters').all()
    for row,g in enumerate(groups,4):
        bd=g.score_breakdown(); wv=g.weighted_final_score()
        vals=[g.name,g.get_presenter_names()]
        for et in ['student','tutor','invited']:
            if et in bd:
                d=bd[et]
                vals+=[d.get('content',0),d.get('presentation_skills',0),
                       d.get('time_management',0),d.get('language',0),
                       d.get('questions_answers',0),d.get('total') or 0,d.get('count',0)]
            else:
                vals+=['']*7
        vals+=[wv or '','f"{round(wv/20*100,1)}%"' if wv else '']
        # fix that f-string above
        if wv:
            vals[-1] = f"{round(wv/20*100,1)}%"
        row_bgs=['FFFFFF','F9FAFB']+['EFF6FF']*7+['F0FDF4']*7+['F5F3FF']*7+['EEF2FF']*2
        for c,v in enumerate(vals,1):
            cell=ws2.cell(row=row,column=c,value=v)
            cell.font=bf if c<=2 else nf; cell.alignment=ctr
            cell.border=bdr; cell.fill=PatternFill('solid',start_color=row_bgs[c-1])

    # ── Sheet 3: Rankings ──
    ws3=wb.create_sheet("Final Rankings"); ws3.freeze_panes='A3'
    ws3.merge_cells('A1:I1')
    t3=ws3['A1']; t3.value="Final Group Rankings"
    t3.font=Font(name='Arial',bold=True,size=13,color='FFFFFF')
    t3.alignment=ctr; t3.fill=PatternFill('solid',start_color=C_HDR)
    ws3.row_dimensions[1].height=30

    rh=['Rank','Group','Presenters','Students Avg (/20)','Tutors Avg (/20)',
        'Invited Avg (/20)','Weighted Final (/20)','% Score','Grade']
    rw=[8,18,28,20,18,20,22,12,10]
    for c,(h,w) in enumerate(zip(rh,rw),1):
        hdr(ws3.cell(row=2,column=c),h,C_FIN)
        ws3.column_dimensions[get_column_letter(c)].width=w
    ws3.row_dimensions[2].height=35

    medals={1:'🥇',2:'🥈',3:'🥉'}
    for r in build_rankings():
        rk=r['rank']; wv=r['weighted_score']
        pct=round(wv/20*100,1) if wv else None
        grade=''
        if pct:
            grade='D' if pct>=17 else 'VGP' if pct>=16 else 'GP' if pct>=14 else 'GP' if pct>=12 else 'P' if pct>=10 else 'G'
        row_bg='FFFBEB' if rk==1 else 'F8FAFC' if rk%2==0 else 'FFFFFF'
        vals=[f"{medals.get(rk,'')} {rk}",r['group'].name,r['presenters'],
              r['student_avg'] or '',r['tutor_avg'] or '',r['invited_avg'] or '',
              wv or '',f"{pct}%" if pct else '',grade]
        for c,v in enumerate(vals,1):
            cell=ws3.cell(row=rk+2,column=c,value=v)
            cell.font=bf if rk<=3 else nf; cell.alignment=ctr
            cell.border=bdr; cell.fill=PatternFill('solid',start_color=row_bg)

    if config:
        nr=len(list(groups))+4
        ws3.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=9)
        nc=ws3.cell(row=nr,column=1)
        nc.value=(f"Config: {config.name}  |  Students: {config.student_weight}%  |  "
                  f"Tutors: {config.tutor_weight}%  |  Invited: {config.invited_weight}%")
        nc.font=Font(name='Arial',italic=True,size=9,color='6B7280'); nc.alignment=ctr

    response=HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition']='attachment; filename="presentation_evaluation_results.xlsx"'
    wb.save(response)
    return response
